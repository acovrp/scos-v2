# SCOS 2.0 Authoritative Product Catalog Compiler
# Location: C:\Users\User\Downloads\scos-v2\pipeline\build_catalog.py

import os
import json
import sqlite3
import pandas as pd
import openpyxl
from database import DB_PATH, get_connection

mrp_excel_path = r"C:\Users\User\Downloads\SKU MRPs as of 19th May.xlsx"
fk_clf_path = r"C:\Users\User\Downloads\sleepycat-agent\sleepycat-agent\agent_data\fk_clf_by_sku_actual.csv"
catalog_json_path = r"C:\Users\User\Downloads\scos-v2\data\catalog.json"
preview_path = r"C:\Users\User\Downloads\scos-v2\data\catalog_preview.txt"

NAME_TO_LINE = {
    'Hybrid Latex Mattress': 'hybridla',
    'Original Mattress': 'original',
    'Tri-fold Mattress': 'trifoldm',
    'Tri Fold Mattress': 'trifoldm',
    'Ultima Mattress': 'ultimama',
    'Ultima Latex Mattress': 'ultimala',
    'Latex Ortho Mattress': 'latexort',
    'The Latex Ortho': 'latexort',
    'The Latex Ortho Mattress': 'latexort',
    'Cloud Spring Mattress': 'cloudspr',
    'Baby Mattress': 'babymatt',
    'Baby Natural Latex Mattress': 'babymatt',
    'Natural Latex Baby Mattress': 'babymatt',
    'Dual Switch Mattress': 'switchdual',
    'Memory Foam Mattress Topper': 'memoryfo',
    'Mattress Topper': 'memoryfo',
    'Topper': 'memoryfo',
    
    # Pillows
    'MF Pillow Bamboo': 'mfpillow',
    'MF Pillow (Bamboo)': 'mfpillow',
    'Cloud Pillow': 'cloudpil',
    'Cloud Pillow - Standard Size': 'cloudpil',
    'SoftTouch Memory Foam Pillow': 'softtouc',
    'Soft Touch Memory Mattress': 'softtouc',
    'SleepyCat Lite Series Microfibre Solid Sleeping Pillow': 'sleepyca',
    'SleepyCat Lite': 'sleepyca',
    'Cuddle Pillow': 'cuddlepi',
    'Pregnancy / Body Pillow': 'cuddlepi',
    'Cervical Bamboo Pillows': 'cervical',
    'Cervical Bamboo Pillow': 'cervical',
    'Contour Bamboo MF Pillow': 'cervical',
    'Contour Latex Pillow': 'contourlatex',
    'Contour CoolTEC Memory Foam Pillow': 'contourcool',
    'CoolTEC MF Pillow': 'cooltec',
    'CoolTEC Memory Foam Pillow': 'cooltec',
    'Dual Comfort Pillow': 'dualcomfort',
    'Fiber Lite': 'fiberlite',
    'Hybrid Pillow': 'hybridpillow',
    'Marshmallow Pillow': 'marshmallow',
    'Travel Neck Pillow': 'travelneck',
    'Wedge Pillow': 'wedge',
    
    # Bedding & Accessories
    'Comforter': 'comforte',
    'Luxe Comforter': 'comforte',
    'Summer Luxe Comforter': 'comforte',
    '150 GSM Summer Luxe Comforter': 'comforte',
    'Winter Comforter': 'comforte',
    'Summer AC Comforter': 'comforte',
    'Bedsheet Set': 'bedshee',
    'Sateen Fitted Sheets': 'bedshee',
    'Jersey Fitted Bed sheet': 'bedshee',
    'Jersey Fitted BedSheet': 'bedshee',
    'Cotton Fitted Bedsheet': 'bedshee',
    'Cotton Fitted BedSheet': 'bedshee',
    'Sateen Duvet Cover': 'bedshee',
    'Duvet Cover': 'bedshee',
    'Cotton Duvet cover': 'bedshee',
    'Jersey Duvet Cover': 'bedshee',
    'Knitted Throw': 'bedshee',
    'Muslin Throws': 'bedshee',
    'Mattress Protector': 'protector',
    'Protector': 'protector',
    'Quilted Mattress Protector': 'protector',
    'Mattress Cover': 'protector',
    
    # Pillow Cases
    'Satin Pillow Case': 'pillowcase',
    'Jersey Pillow Cases': 'pillowcase',
    'Cotton Pillow Case': 'pillowcase',
    'Sateen Pillow Case': 'pillowcase',
    'Cuddle Pillow Case': 'pillowcase',
    
    # Pets
    'Pet Bed': 'dogbed',
    'Dog Bed - Original': 'dogbed',
    'Dog Bed - Orthopedic': 'dogbed',
    
    # Furniture & Others
    'Enso Recliner': 'recliner',
    'Seika Recliner': 'recliner',
    'Recliner': 'recliner',
    'Katachi Bed': 'bed',
    'Ohayo Bed': 'bed',
    'Bed': 'bed',
    'Taurus Smart Recliner Bed With Headboard': 'bed'
}

def get_parent_line(name):
    name = name.strip()
    if name in NAME_TO_LINE:
        return NAME_TO_LINE[name]
    
    # Substring checks
    for key, line in NAME_TO_LINE.items():
        if key.lower() in name.lower() or name.lower() in key.lower():
            return line
            
    if "mattress" in name.lower():
        return "original"
    if "pillow" in name.lower():
        return "cloudpil"
    if "bed" in name.lower():
        return "bed"
    return "other"

def normalize_category(category_name):
    if not category_name:
        return 'Other'
    c_lower = str(category_name).lower().strip()
    if 'mattress' in c_lower:
        return 'Mattress'
    if 'pillow' in c_lower:
        return 'Pillow'
    if 'bedding' in c_lower or 'sheets' in c_lower or 'duvet' in c_lower or 'throw' in c_lower or 'case' in c_lower:
        return 'Bedding'
    if 'topper' in c_lower:
        return 'Topper'
    if 'bed' in c_lower or 'recliner' in c_lower or 'furniture' in c_lower:
        return 'Furniture'
    return 'Other'

def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ('#N/A', 'None', 'none', 'null', 'NULL', ''):
        return None
    return s

def build_catalog():
    print("Building canonical catalog from Excel master...")
    
    # 1. Load FSN lookup from Flipkart CLF actuals
    fsn_lookup = {}
    if os.path.exists(fk_clf_path):
        df_clf = pd.read_csv(fk_clf_path)
        for _, row in df_clf.iterrows():
            sku = clean_str(row.get('sku'))
            fsn = clean_str(row.get('fsn'))
            if sku and fsn:
                fsn_lookup[sku] = fsn
    else:
        print(f"Warning: CLF file not found at {fk_clf_path}. FSN mappings will be skipped.")

    # 2. Parse Excel master
    if not os.path.exists(mrp_excel_path):
        raise FileNotFoundError(f"Authoritative Excel master not found at {mrp_excel_path}")
        
    wb = openpyxl.load_workbook(mrp_excel_path, read_only=True, data_only=True)
    ws = wb['New Master']
    
    headers = [clean_str(cell.value) for cell in next(ws.iter_rows())]
    
    idx_master_sku = headers.index('Master SKU')
    idx_asin = headers.index('ASIN')
    idx_flex_status = headers.index('Flex Status') if 'Flex Status' in headers else -1
    idx_flipkart_sku = headers.index('Flipkart')
    idx_product = headers.index('PRODUCT')
    idx_category = headers.index('CATEGORY')
    idx_mrp = headers.index('MRP')
    
    unified_products = {}
    skipped_rows = 0
    duplicate_rows = 0
    
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        master_sku = clean_str(row_cells[idx_master_sku])
        if not master_sku:
            skipped_rows += 1
            continue
            
        asin = clean_str(row_cells[idx_asin])
        fk_sku = clean_str(row_cells[idx_flipkart_sku])
        prod_name = clean_str(row_cells[idx_product]) or "Unknown Product"
        category_raw = clean_str(row_cells[idx_category]) or "Other"
        
        mrp_raw = row_cells[idx_mrp]
        try:
            mrp = float(mrp_raw) if mrp_raw is not None else None
        except Exception:
            mrp = None
            
        flex_status = clean_str(row_cells[idx_flex_status]) if idx_flex_status != -1 else None
        category = normalize_category(category_raw)
        line = get_parent_line(prod_name)
        fsn = fsn_lookup.get(fk_sku) if fk_sku else None
        
        # Key by variant signature to allow multiple ASINs/Flipkart SKUs for the same Master SKU
        key = (master_sku, asin, fk_sku)
        if key in unified_products:
            duplicate_rows += 1
            existing = unified_products[key]
            if flex_status == 'active' and existing.get('flex_status') != 'active':
                # Overwrite existing with active row
                pass
            else:
                # Keep existing, skip current
                continue
                
        unified_products[key] = {
            'sku': master_sku,
            'asin': asin,
            'fk_sku': fk_sku,
            'fsn': fsn,
            'name': prod_name,
            'category': category,
            'line': line,
            'mrp': mrp,
            'flex_status': flex_status
        }
        
    # Define additional mappings for 100% coverage
    ADDITIONAL_PRODUCTS = [
        # Amazon ASINs
        {
            'sku': 'SC-A-SLMMFPILWSET1-BAM-S-25x15',
            'asin': 'B0DZP8GXDV',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Slim Memory Foam Pillow',
            'category': 'Pillow',
            'line': 'cloudpil',
            'mrp': 1999.0,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-LITEPILWPSET4-S-27X17',
            'asin': 'B0GLNNJ51N',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Fiber Lite Pillow',
            'category': 'Pillow',
            'line': 'cloudpil',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-SLMMFPILWSET2-BAM-S-25x15',
            'asin': 'B0GQ9TRGRT',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Slim Memory Foam Pillow',
            'category': 'Pillow',
            'line': 'cloudpil',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-CON-MFPILWS2-CLT-25x16',
            'asin': 'B0GYF7QZ8Q',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat CoolTEC Cervical Memory Foam Pillow',
            'category': 'Pillow',
            'line': 'cervical',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-AF-PETBEDORIG-N.Blue-XL-54x36',
            'asin': 'B0FSTJ7FFJ',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Original Dog Bed',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-AF-PETBEDORIG-M.Red-L-46x32',
            'asin': 'B0FSTCNH49',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Original Dog Bed',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-NLX-BABY-L-48x24x4',
            'asin': 'B0GVG2KQDX',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Natural Latex Baby Mattress',
            'category': 'Mattress',
            'line': 'original',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-SLMPILWSET4-S-26x16',
            'asin': 'B0GLNFQTY9',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Slim Pillow with Microfiber',
            'category': 'Pillow',
            'line': 'cloudpil',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-Ohayo-Double-Bed-78x48',
            'asin': 'B0FR9CYKW6',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Ohayo Bed',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-AF-PETBEDORTHO-N.Blue-M-28x26',
            'asin': 'B0FSTD9JDC',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Orthopedic Dog Bed',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-Ohayo-King-Bed-72x72',
            'asin': 'B0FR98QN94',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Ohayo Bed',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-Ohayo-King-Bed-75x72',
            'asin': 'B0FR98NSZ9',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Ohayo Bed',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-NLX-BABY-XL-52x28x4',
            'asin': 'B0GVFH6X8X',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Natural Latex Baby Mattress',
            'category': 'Mattress',
            'line': 'original',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-Ohayo-Queen-Bed-75x60',
            'asin': 'B0FR97MMHL',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Ohayo Bed',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-ASF-Ohayo-Queen-Bed-72x60',
            'asin': 'B0FR96YX38',
            'fk_sku': None,
            'fsn': None,
            'name': 'SleepyCat Ohayo Bed',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        # Flipkart SKUs
        {
            'sku': 'SCORIG-78x72x6',
            'asin': None,
            'fk_sku': 'SCORIG-78x72x6',
            'fsn': 'BEMHER3NY9ENMDVZ',
            'name': 'Original Mattress 78x72x6',
            'category': 'Mattress',
            'line': 'original',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-Fk-ORIG-D-72x48x5',
            'asin': None,
            'fk_sku': 'SC-Fk-ORIG-D-72x48x5',
            'fsn': 'BEMHHU8GHVCABY9M',
            'name': 'Original Mattress Double 72x48x5',
            'category': 'Mattress',
            'line': 'original',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-F-MFPILW-P-28x18',
            'asin': None,
            'fk_sku': 'SC-F-MFPILW-P-28x18',
            'fsn': 'PLWG3UXUY6ZDARH7',
            'name': 'Memory Foam Pillow 28x18',
            'category': 'Pillow',
            'line': 'mfpillow',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-PETBEDORTHO-M.Red-XL-52X28',
            'asin': None,
            'fk_sku': 'SC-PETBEDORTHO-M.Red-XL-52X28',
            'fsn': 'PEBGF39RXZFGZKEK',
            'name': 'Orthopedic Dog Bed Red XL',
            'category': 'Other',
            'line': 'protector',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-FK-CNC-BLK-SMALL',
            'asin': None,
            'fk_sku': 'SC-FK-CNC-BLK-SMALL',
            'fsn': None,
            'name': 'Cuddle Pillow Small Black',
            'category': 'Pillow',
            'line': 'cuddlepi',
            'mrp': None,
            'flex_status': 'active'
        },
        {
            'sku': 'SC-F-SLSTPILWCS-C-52x20',
            'asin': None,
            'fk_sku': 'SC-F-SLSTPILWCS-C-52x20',
            'fsn': None,
            'name': 'SoftTouch Sleep Pillow Case 52x20',
            'category': 'Bedding',
            'line': 'bedshee',
            'mrp': None,
            'flex_status': 'active'
        }
    ]

    # Add additional items to unified_products
    for item in ADDITIONAL_PRODUCTS:
        key = (item['sku'], item['asin'], item['fk_sku'])
        unified_products[key] = item

    print(f"Parsed {len(unified_products)} unique product variants (including manually mapped).")
    print(f"Skipped {skipped_rows} empty rows, resolved {duplicate_rows} duplicate variant rows.")
    
    # 3. Seed SQLite products table
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM products")
    
    # Insert variants
    for item in unified_products.values():
        cursor.execute("""
            INSERT INTO products (sku, asin, fk_sku, fsn, name, category, line, mrp)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (item['sku'], item['asin'], item['fk_sku'], item['fsn'], item['name'], item['category'], item['line'], item['mrp']))
        
    conn.commit()
    conn.close()
    print(f"Seeded products table in SQLite Core database.")
    
    # Convert tuple keys to unique string keys for JSON serialization
    serializable_catalog = {}
    for idx, (k, item) in enumerate(unified_products.items()):
        str_key = item['sku']
        if str_key in serializable_catalog:
            str_key = f"{item['sku']}_{idx}"
        serializable_catalog[str_key] = item

    # 4. Generate catalog config in both JSON and JS format (JSON for APIs, JS for direct file:// load)
    os.makedirs(os.path.dirname(catalog_json_path), exist_ok=True)
    with open(catalog_json_path, 'w', encoding='utf-8') as f:
        json.dump(serializable_catalog, f, indent=2, ensure_ascii=False)
    
    catalog_js_path = catalog_json_path.replace(".json", ".js")
    with open(catalog_js_path, 'w', encoding='utf-8') as f:
        f.write(f"const CATALOG_DATA = {json.dumps(serializable_catalog, indent=2, ensure_ascii=False)};")
    print(f"Exported SCOS 2.0 authoritative catalog to {catalog_json_path} and {catalog_js_path}")
    
    # 5. Write visual summary report to data/catalog_preview.txt
    with open(preview_path, 'w', encoding='utf-8') as f:
        f.write("=========================================\n")
        f.write("SCOS 2.0 DEDUPLICATED PRODUCT CATALOG PREVIEW\n")
        f.write("=========================================\n")
        f.write(f"Source file: {mrp_excel_path}\n")
        f.write(f"Total Unique Product Variants: {len(unified_products)}\n")
        f.write("=========================================\n\n")
        
        # Group by category
        grouped = {}
        for p in unified_products.values():
            cat = p['category']
            if cat not in grouped:
                grouped[cat] = []
            grouped[cat].append(p)
            
        for cat in sorted(grouped.keys()):
            f.write(f"CATEGORY: {cat}\n")
            f.write("-" * 60 + "\n")
            for item in sorted(grouped[cat], key=lambda x: (x['line'], x['sku'])):
                f.write(f"  SKU: {item['sku']:32} | Line: {item['line']:12} | ASIN: {str(item['asin']):12} | FK SKU: {str(item['fk_sku']):24} | FSN: {str(item['fsn']):18} | MRP: {str(item['mrp']):8} | Name: {item['name']}\n")
            f.write("\n")
            
    print(f"Catalog summary written to: {preview_path}")

if __name__ == "__main__":
    build_catalog()
